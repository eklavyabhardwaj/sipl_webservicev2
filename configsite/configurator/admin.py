from django.utils.html import format_html
from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import redirect, render
from django import forms
from django.db import transaction
from django.http import HttpResponse
from .models import (
    ProductGroup,
    Item,
    ItemImage,
    Question,
    Choice,
    ChoiceImpact,
    QuizSession,
    Answer,
    ItemSpec,
    ItemDocument,
    ItemFeature, ItemVariantImage,
ItemVariantSpec, ItemVariantDocument, ItemVariant
)
from .models import Page,ERPSettings,ContactMessage
from django import forms
from .models import Question, Choice

class QuestionAdminForm(forms.ModelForm):
    class Meta:
        model = Question
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Prefer POST -> INITIAL -> INSTANCE; handle both pk and object
        parent = (
            self.data.get("depends_on")
            or self.initial.get("depends_on")
            or getattr(self.instance, "depends_on", None)
        )
        # normalize to id
        if hasattr(parent, "pk"):
            parent_id = parent.pk
        else:
            try:
                parent_id = int(parent) if parent is not None else None
            except (TypeError, ValueError):
                parent_id = None

        if "trigger_choices" in self.fields:
            qs = Choice.objects.filter(question_id=parent_id) if parent_id else Choice.objects.none()
            self.fields["trigger_choices"].queryset = qs
            self.fields["trigger_choices"].help_text = (
                "Pick choice(s) from the selected parent question."
                if parent_id else
                "Select a parent in 'Depends on' and Save to load choices."
            )


class ItemVariantImageInline(admin.TabularInline):
    model = ItemVariantImage
    extra = 1
    fields = ("image", "alt_text", "created_at")
    readonly_fields = ("created_at",)

class ItemVariantSpecInline(admin.TabularInline):
    model = ItemVariantSpec
    extra = 3
    fields = ("order", "label", "value", "unit", "highlight")
    ordering = ("order", "id")

class ItemVariantDocumentInline(admin.TabularInline):
    model = ItemVariantDocument
    extra = 1

class ItemVariantInline(admin.StackedInline):
    model = ItemVariant
    extra = 1
    show_change_link = True
    fields = ("name", "code", "is_active", "description")
    inlines = []  # Django Admin doesn't nest inlines; variant children are edited on variant’s change page.


@admin.register(ItemVariant)
class ItemVariantAdmin(admin.ModelAdmin):
    list_display = ("name", "item", "code", "is_active")
    list_filter = ("is_active", "item__group")
    search_fields = ("name", "code", "item__name")
    autocomplete_fields = ("item",)
    inlines = [ItemVariantImageInline, ItemVariantSpecInline, ItemVariantDocumentInline]







@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "phone", "subject", "created_at", "handled")
    list_filter = ("handled", "created_at")
    search_fields = ("name", "email", "phone", "subject", "message")





@admin.register(Page)
class PageAdmin(admin.ModelAdmin):
    list_display = ("title", "slug", "is_active", "is_home", "menu_order", "hero_thumb", "is_external")
    list_filter = ("is_active", "is_home")
    search_fields = ("title", "slug", "body")
    prepopulated_fields = {"slug": ("title",)}
    fields = (
        "title", "slug", "is_active", "is_home", "menu_order",
        "external_url",
        "show_hero", "hero_image", "hero_alt", "hero_caption", "hero_thumb",
        "body", "footer_html",
        "custom_css",
    )
    readonly_fields = ("hero_thumb",)

    def hero_thumb(self, obj):
        if obj.hero_image:
            return format_html('<img src="{}" style="height:60px;border-radius:8px;object-fit:cover;">', obj.hero_image.url)
        return "—"
    hero_thumb.short_description = "Hero preview"

    def is_external(self, obj):
        return bool(obj.external_url)
    is_external.boolean = True
    is_external.short_description = "External?"




class ItemFeatureInline(admin.TabularInline):
    model = ItemFeature
    extra = 2  # number of empty rows to show
    fields = ("text",)




class ItemImageInline(admin.TabularInline):
    model = ItemImage
    extra = 1
    fields = ("preview", "image", "alt_text", "created_at")
    readonly_fields = ("preview", "created_at")

    def preview(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:60px;border-radius:6px;object-fit:cover;">',
                obj.image.url,
            )
        return "—"
    preview.short_description = "Preview"


class ItemDocumentInline(admin.TabularInline):   # NEW
    model = ItemDocument
    extra = 1

class ChoiceImpactInline(admin.TabularInline):
    model = ChoiceImpact
    extra = 1
    autocomplete_fields = ("item",)
    fields = ("item", "score")


class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 2
    show_change_link = True
    fields = ("order", "text", "is_active", "image", "img_thumb", "image_width", "image_height")
    readonly_fields = ("img_thumb", "image_width", "image_height")

    def img_thumb(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:46px;border-radius:6px;object-fit:contain;background:#fff;">',
                obj.image.url,
            )
        return "—"
    img_thumb.short_description = "Image"


class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0
    can_delete = False
    readonly_fields = ("question", "choice")
    fields = ("question", "choice")


# =========================
# ProductGroup
# =========================

@admin.register(ProductGroup)
class ProductGroupAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "is_active", "hero_thumb")
    list_filter = ("is_active",)
    search_fields = ("name", "slug")
    fields = ("name", "slug", "is_active", "hero_image", "hero_thumb")
    readonly_fields = ("hero_thumb",)
    prepopulated_fields = {"slug": ("name",)}

    def hero_thumb(self, obj):
        if obj.hero_image:
            return format_html(
                '<img src="{}" style="height:60px;border-radius:8px;object-fit:cover;">',
                obj.hero_image.url,
            )
        return "—"
    hero_thumb.short_description = "Hero preview"


class ItemSpecInline(admin.TabularInline):
    model = ItemSpec
    extra = 3
    fields = ("order", "label", "value", "unit", "highlight")
    ordering = ("order", "id")



# =========================
# Item
# =========================



@admin.register(Item)
class ItemAdmin(admin.ModelAdmin):
    change_list_template = "admin/configurator/item/change_list.html"
    list_display = ("name", "group", "item_code", "is_active")
    list_filter = ("group", "is_active")
    search_fields = ("name", "item_code", "features__text")  # search into features
    autocomplete_fields = ("group",)
    inlines = [ItemImageInline, ItemSpecInline, ItemDocumentInline, ItemFeatureInline, ItemVariantInline]

    fields = ("group", "name", "item_code", "description", "is_active")

    def changelist_view(self, request, extra_context=None):  # <-- add this
        extra_context = extra_context or {}
        extra_context["import_url"] = "admin:configurator_item_import"  # matches your get_urls() name
        return super().changelist_view(request, extra_context=extra_context)


    # ---- Custom import view ----
    class ImportForm(forms.Form):
        file = forms.FileField(
            help_text="Upload a .csv or .xlsx file. Columns: group_name, item_name, item_code, description, is_active, features, specs"
        )
        mode = forms.ChoiceField(
            choices=[("upsert", "Upsert (create or update)"),
                     ("create", "Create only"),
                     ("update", "Update only")],
            initial="upsert",
        )
        clear_features = forms.BooleanField(
            required=False,
            help_text="If checked, replace existing features with the uploaded list for each item."
        )
        feature_separator = forms.CharField(
            required=False,
            initial=";",
            help_text="Separator for multiple features in a single cell (default ';'). Newlines also work."
        )
        # NEW ↓
        clear_specs = forms.BooleanField(
            required=False,
            help_text="If checked, replace existing specs with the uploaded list for each item."
        )

    def _parse_specs(self, raw: str):
        """
        Parse 'specs' column into a list of dicts:
          Required: label, value
          Optional: unit, order, highlight
        Format per spec (pipe-separated k=v), multiple specs separated by ';'
          Example: label=CPU|value=Intel i5|unit=—|order=1|highlight=0
        """
        specs = []
        raw = (raw or "").strip()
        if not raw:
            return specs

        items = [p.strip() for p in raw.split(";") if p.strip()]
        for it in items:
            parts = [p.strip() for p in it.split("|") if p.strip()]
            kv = {}
            for part in parts:
                if "=" in part:
                    k, v = part.split("=", 1)
                    kv[k.strip().lower()] = v.strip()
            if "label" in kv and "value" in kv:
                specs.append(kv)
        return specs

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("import/", self.admin_site.admin_view(self.import_items), name="configurator_item_import"),
            path("import/template/", self.admin_site.admin_view(self.items_template_csv),
                 name="configurator_item_template"),
        ]
        return my_urls + urls

    def items_template_csv(self, request):
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="items_import_template.csv"'

        writer = csv.writer(response)
        # header — add "specs"
        writer.writerow(["group_name", "item_name", "item_code", "description", "is_active", "features", "specs"])

        # sample rows with a specs example
        writer.writerow([
            "Laptops", "ProBook 14", "LPB-014", "Lightweight business laptop", "true",
            "14-inch display;Intel i5;8GB RAM;512GB SSD;Backlit keyboard",
            "label=CPU|value=Intel i5; label=RAM|value=8GB|highlight=1; label=Storage|value=512|unit=GB SSD|order=3"
        ])
        writer.writerow([
            "Laptops", "UltraSlim 13", "ULS-013", "Thin-and-light ultraportable", "1",
            "13.3-inch display;Intel i7;16GB RAM;1TB SSD;Thunderbolt 4",
            "label=CPU|value=Intel i7; label=RAM|value=16GB; label=Storage|value=1|unit=TB SSD"
        ])
        writer.writerow([
            "Desktops", "Workstation X", "WSX-001", "High-performance desktop workstation", "true",
            "Ryzen 9;64GB RAM;RTX 4070;2TB NVMe;ECC support",
            "label=CPU|value=Ryzen 9; label=RAM|value=64GB; label=GPU|value=RTX 4070; label=Storage|value=2|unit=TB NVMe; label=ECC|value=Yes|highlight=1"
        ])
        return response

    def import_items(self, request):
        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import Items from CSV/XLSX",
        }




        if request.method == "POST":
            form = self.ImportForm(request.POST, request.FILES)
            if form.is_valid():
                f = form.cleaned_data["file"]
                mode = form.cleaned_data["mode"]
                clear_features = form.cleaned_data["clear_features"]
                sep = (form.cleaned_data["feature_separator"] or ";").strip()
                clear_specs = form.cleaned_data.get("clear_specs", False)

                try:
                    rows = self._read_rows(f)
                except Exception as e:
                    messages.error(request, f"Could not read file: {e}")
                    context["form"] = form
                    return render(request, "admin/import_items.html", context)

                created, updated, skipped = 0, 0, 0
                errors = []

                with transaction.atomic():
                    for i, row in enumerate(rows, start=2):  # assume header at row 1
                        try:
                            group_name = (row.get("group_name") or "").strip()
                            item_name  = (row.get("item_name") or "").strip()
                            item_code  = (row.get("item_code") or "").strip() or None
                            description = (row.get("description") or "").strip()
                            is_active = str(row.get("is_active") or "").strip().lower()

                            # interpret is_active
                            is_active_val = True
                            if is_active in ("", "true", "1", "yes", "y"):
                                is_active_val = True
                            elif is_active in ("false", "0", "no", "n"):
                                is_active_val = False

                            features_raw = row.get("features") or ""
                            features_list = self._split_features(features_raw, sep)

                            specs_raw = (row.get("specs") or "").strip()
                            specs_list = self._parse_specs(specs_raw)

                            if not group_name or not item_name:
                                skipped += 1
                                continue

                            group, _ = ProductGroup.objects.get_or_create(name=group_name)

                            # Matching priority: item_code (if provided), else (group, name)
                            item_qs = Item.objects.all()
                            item_obj = None
                            if item_code:
                                item_obj = item_qs.filter(item_code=item_code).first()
                            if not item_obj:
                                item_obj = item_qs.filter(group=group, name=item_name).first()

                            if item_obj:
                                if mode == "create":
                                    skipped += 1
                                    continue
                                # update
                                item_obj.name = item_name
                                item_obj.group = group
                                item_obj.description = description
                                item_obj.is_active = is_active_val
                                if item_code:
                                    item_obj.item_code = item_code
                                item_obj.save()
                                updated += 1
                            else:
                                if mode == "update":
                                    skipped += 1
                                    continue
                                # create
                                item_obj = Item.objects.create(
                                    group=group,
                                    name=item_name,
                                    item_code=item_code,
                                    description=description,
                                    is_active=is_active_val,
                                )
                                created += 1

                            # features
                            if clear_features:
                                item_obj.features.all().delete()

                            for ft in features_list:
                                if ft:
                                    ItemFeature.objects.get_or_create(item=item_obj, text=ft)

                            # NEW: specs (uses related_name="specs" on ItemSpec)
                            if clear_specs:
                                item_obj.specs.all().delete()  # ItemSpec FK uses related_name="specs"

                            for sp in specs_list:
                                lbl = sp.get("label")
                                val = sp.get("value", "")
                                if not lbl:
                                    continue
                                spec_obj, _ = ItemSpec.objects.get_or_create(item=item_obj, label=lbl)
                                spec_obj.value = val
                                if "unit" in sp:
                                    spec_obj.unit = sp["unit"]
                                if "order" in sp:
                                    try:
                                        spec_obj.order = int(sp["order"])
                                    except Exception:
                                        pass
                                if "highlight" in sp:
                                    spec_obj.highlight = str(sp["highlight"]).strip().lower() in ("1", "true", "yes",
                                                                                                  "y")
                                spec_obj.save()

                        except Exception as e:
                            errors.append(f"Row {i}: {e}")

                if errors:
                    for e in errors:
                        messages.error(request, e)
                messages.success(request, f"Import complete — created: {created}, updated: {updated}, skipped: {skipped}")
                return redirect("admin:configurator_item_changelist")
        else:
            form = self.ImportForm()

        context["form"] = form
        # Fallback minimal template if you don't want to create a file:
        # We render a very small form inline.
        return render(request, "admin/import_items.html", context)

    # --- helpers ---
    def _read_rows(self, uploaded_file):
        name = (uploaded_file.name or "").lower()
        if name.endswith(".csv"):
            import csv, io
            text = uploaded_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        elif name.endswith(".xlsx"):
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl is required for .xlsx files. Install with: pip install openpyxl")
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[i]: ("" if val is None else str(val)) for i, val in enumerate(r)}
                rows.append(row)
            return rows
        else:
            raise RuntimeError("Unsupported file type. Please upload .csv or .xlsx")

    def _split_features(self, raw: str, sep: str):
        import re
        raw = str(raw or "")
        parts = []
        # split on chosen separator or newlines
        if sep:
            parts.extend([p.strip() for p in str(raw).split(sep)])
        # also split on newlines
        nl_parts = re.split(r"[\r\n]+", raw)
        for p in nl_parts:
            p = p.strip()
            if p and p not in parts:
                parts.append(p)
        # dedupe empties
        return [p for p in parts if p]




# =========================
# Question
# =========================

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    form = QuestionAdminForm  # <-- add this line

    list_display = ("text", "question_tag", "group", "order", "input_type",
                    "affects_score", "is_required", "is_active", "depends_on_display", "img_thumb")
    list_filter = ("group", "is_active", "is_required", "input_type", "affects_score")
    search_fields = ("text", "group__name")
    ordering = ("group", "order", "id")
    inlines = [ChoiceInline]
    autocomplete_fields = ("group",)  # you can also add "depends_on" if desired
    filter_horizontal = ("trigger_choices",)  # nicer M2M UI  # <-- helps with M2M UI

    fieldsets = (
        (None, {
            "fields": ("group", "text", "question_tag", "order", "is_required", "is_active")
        }),
        ("Behavior", {
            "fields": ("input_type", "affects_score")
        }),
        ("Conditional visibility", {  # <-- show your new fields here
            "fields": ("depends_on", "trigger_choices"),
            "description": "Show this question only if the selected parent’s specific choice(s) are chosen."
        }),
        ("Image", {
            "fields": ("image", "img_thumb", "image_width", "image_height")
        }),
    )
    readonly_fields = ("img_thumb", "image_width", "image_height")

    def depends_on_display(self, obj):
        return obj.depends_on.text if obj.depends_on_id else "—"
    depends_on_display.short_description = "Depends on"


    def img_thumb(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:80px;border-radius:8px;object-fit:contain;background:#fff;">',
                obj.image.url,
            )
        return "—"
    img_thumb.short_description = "Image"

    # --- add inside your existing QuestionAdmin ---

    change_list_template = "admin/configurator/question/change_list.html"

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        # name used in the template url tag
        extra_context["import_url"] = "admin:configurator_question_import"
        return super().changelist_view(request, extra_context=extra_context)

    # -------- Import form for Questions --------
    class QuestionImportForm(forms.Form):
        file = forms.FileField(
            help_text=(
                "Upload .csv or .xlsx with columns: "
                "group_name, text, input_type, choices, is_required, is_active, affects_score, order, question_tag"
            )
        )
        mode = forms.ChoiceField(
            choices=[
                ("upsert", "Upsert (create or update)"),
                ("create", "Create only"),
                ("update", "Update only"),
            ],
            initial="upsert",
        )
        clear_choices = forms.BooleanField(
            required=False,
            help_text="If checked, delete existing choices for the question before inserting uploaded choices."
        )
        choices_separator = forms.CharField(
            required=False,
            initial=";",
            help_text="Separator for multiple choices (default ';'). Newlines also work."
        )

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_questions),
                name="configurator_question_import",
            ),
            path(
                "import/template/",
                self.admin_site.admin_view(self.questions_template_csv),
                name="configurator_question_template",
            ),
        ]
        return my_urls + urls

    def questions_template_csv(self, request):
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="questions_import_template.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "group_name", "text", "input_type", "choices",
            "is_required", "is_active", "affects_score",
            "order", "question_tag"
        ])
        # Sample rows
        writer.writerow([
            "Laptops",
            "What screen size do you prefer?",
            "single",
            "label=13-inch|order=1|active=1; label=14-inch|order=2|active=1; label=15-inch|order=3|active=1",
            "1", "1", "1", "10", "screen_size"
        ])
        writer.writerow([
            "Laptops",
            "Pick the features you care about",
            "multi",
            "label=Backlit keyboard|order=1|active=1; label=Touchscreen|order=2|active=1; label=Fingerprint|order=3|active=1",
            "0", "1", "0", "20", "features"
        ])
        return response

    def import_questions(self, request):
        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import Questions from CSV/XLSX",
        }

        if request.method == "POST":
            form = self.QuestionImportForm(request.POST, request.FILES)
            if form.is_valid():
                uploaded = form.cleaned_data["file"]
                mode = form.cleaned_data["mode"]
                clear_choices = form.cleaned_data["clear_choices"]
                sep = (form.cleaned_data["choices_separator"] or ";").strip()

                try:
                    rows = self._read_rows(uploaded)
                except Exception as e:
                    messages.error(request, f"Could not read file: {e}")
                    context["form"] = form
                    return render(request, "admin/import_questions.html", context)

                created, updated, skipped = 0, 0, 0
                errors = []

                from .models import ProductGroup, Question, Choice  # local import to avoid cycles

                with transaction.atomic():
                    for i, row in enumerate(rows, start=2):  # header at row 1
                        try:
                            group_name = (row.get("group_name") or "").strip()
                            text = (row.get("text") or "").strip()
                            input_type_raw = (row.get("input_type") or "").strip().lower()
                            choices_raw = row.get("choices") or ""
                            is_required = self._parse_bool(row.get("is_required"))
                            is_active = self._parse_bool(row.get("is_active"))
                            affects_score = self._parse_bool(row.get("affects_score"))
                            order_val = self._parse_int(row.get("order"), default=0)
                            question_tag = (row.get("question_tag") or "").strip()

                            if not group_name or not text:
                                skipped += 1
                                continue

                            group, _ = ProductGroup.objects.get_or_create(name=group_name)

                            # normalize input_type
                            if input_type_raw in ("multi", "multiple", "checkbox", "checkboxes"):
                                input_type_val = Question.INPUT_MULTI
                            else:
                                # default single (radio)
                                input_type_val = Question.INPUT_SINGLE

                            # Find existing question by (group, text)
                            q = Question.objects.filter(group=group, text=text).first()

                            if q:
                                if mode == "create":
                                    skipped += 1
                                    continue
                                # update
                                q.input_type = input_type_val
                                q.is_required = is_required
                                q.is_active = is_active
                                q.affects_score = affects_score
                                q.order = order_val
                                q.question_tag = question_tag
                                q.save()
                                updated += 1
                            else:
                                if mode == "update":
                                    skipped += 1
                                    continue
                                # create
                                q = Question.objects.create(
                                    group=group,
                                    text=text,
                                    input_type=input_type_val,
                                    is_required=is_required,
                                    is_active=is_active,
                                    affects_score=affects_score,
                                    order=order_val,
                                    question_tag=question_tag,
                                )
                                created += 1

                            # Choices
                            choice_specs = self._parse_choices(choices_raw, sep)
                            if clear_choices:
                                q.choices.all().delete()

                            for spec in choice_specs:
                                label = spec.get("label", "").strip()
                                if not label:
                                    continue
                                ch, _ = Choice.objects.get_or_create(question=q, text=label)
                                ch.order = self._parse_int(spec.get("order"), default=0)
                                ch.is_active = self._parse_bool(spec.get("active"))
                                ch.save()

                        except Exception as e:
                            errors.append(f"Row {i}: {e}")

                if errors:
                    for e in errors:
                        messages.error(request, e)
                messages.success(
                    request,
                    f"Question import complete — created: {created}, updated: {updated}, skipped: {skipped}"
                )
                return redirect("admin:configurator_question_changelist")
        else:
            form = self.QuestionImportForm()

        context["form"] = form
        return render(request, "admin/import_questions.html", context)

    # --- helpers (reuse pattern from Item import) ---
    def _read_rows(self, uploaded_file):
        name = (uploaded_file.name or "").lower()
        if name.endswith(".csv"):
            import csv, io
            text = uploaded_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        elif name.endswith(".xlsx"):
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl is required for .xlsx files. Install with: pip install openpyxl")
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in
                       next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[i]: ("" if val is None else str(val)) for i, val in enumerate(r)}
                rows.append(row)
            return rows
        else:
            raise RuntimeError("Unsupported file type. Please upload .csv or .xlsx")

    def _parse_bool(self, val, default=False):
        s = str(val or "").strip().lower()
        if s in ("1", "true", "yes", "y"):
            return True
        if s in ("0", "false", "no", "n"):
            return False
        return default

    def _parse_int(self, val, default=0):
        try:
            return int(str(val).strip())
        except Exception:
            return default

    def _parse_choices(self, raw: str, sep: str):
        """
        Parses 'choices' column into list of dicts.
        Each choice spec is separated by `sep` (default ';') or newline.
        Each spec is pipe-separated key=val pairs. Supported keys:
          - label (required)
          - order (int, optional)
          - active (bool, optional)
        Example cell:
          label=13-inch|order=1|active=1; label=14-inch|order=2|active=1
        """
        raw = (raw or "").strip()
        if not raw:
            return []
        # split on separator and newlines
        import re
        chunks = []
        if sep:
            chunks.extend([p.strip() for p in raw.split(sep)])
        for p in re.split(r"[\r\n]+", raw):
            p = p.strip()
            if p and p not in chunks:
                chunks.append(p)

        out = []
        for ch in chunks:
            if not ch:
                continue
            spec = {}
            parts = [p.strip() for p in ch.split("|") if p.strip()]
            for part in parts:
                if "=" in part:
                    k, v = part.split("=", 1)
                    spec[k.strip().lower()] = v.strip()
                else:
                    # allow bare label fallback
                    spec.setdefault("label", part.strip())
            if "label" in spec:
                out.append(spec)
        return out



# =========================
# Choice
# =========================

@admin.register(Choice)
class ChoiceAdmin(admin.ModelAdmin):
    list_display = ("text", "question", "order", "is_active", "img_thumb")
    list_filter = ("question__group", "is_active")
    search_fields = ("text", "question__text", "question__group__name")
    ordering = ("question", "order", "id")
    inlines = [ChoiceImpactInline]
    autocomplete_fields = ("question",)

    fields = ("question", "order", "text", "is_active", "image", "img_thumb", "image_width", "image_height")
    readonly_fields = ("img_thumb", "image_width", "image_height")

    def img_thumb(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:60px;border-radius:6px;object-fit:contain;background:#fff;">',
                obj.image.url,
            )
        return "—"
    img_thumb.short_description = "Image"


# =========================
# QuizSession
# =========================

@admin.register(QuizSession)
class QuizSessionAdmin(admin.ModelAdmin):
    list_display = ("id", "group", "created_at", "recommended_item", "name", "email", "phone", "company")
    list_filter = ("group", "created_at")
    search_fields = ("name", "email", "phone", "company")
    readonly_fields = ("created_at",)
    filter_horizontal = ("interested_items",)
    autocomplete_fields = ("group", "recommended_item")
    inlines = [AnswerInline]


# =========================
# ChoiceImpact (optional separate view)
# =========================

@admin.register(ChoiceImpact)
class ChoiceImpactAdmin(admin.ModelAdmin):
    list_display = ("choice", "item", "score")
    list_filter = ("choice__question__group",)
    search_fields = ("choice__text", "item__name")
    autocomplete_fields = ("choice", "item")


# =========================
# (Optional) Answer direct view
# =========================
# You generally don’t need to expose Answer directly since it’s inline on QuizSession.
# But if you want it visible/searchable, uncomment below:
#
# @admin.register(Answer)
# class AnswerAdmin(admin.ModelAdmin):
#     list_display = ("session", "question", "choice")
#     list_filter = ("session__group",)
#     search_fields = ("session__name", "question__text", "choice__text")
#     autocomplete_fields = ("session", "question", "choice")
class ERPSettingsForm(forms.ModelForm):
    class Meta:
        model = ERPSettings
        fields = "__all__"
        widgets = {
            "api_key": forms.PasswordInput(render_value=True),
            "api_secret": forms.PasswordInput(render_value=True),
        }


@admin.register(ERPSettings)
class ERPSettingsAdmin(admin.ModelAdmin):
    form = ERPSettingsForm  # <- add this
    list_display = (
        "is_enabled",
        "base_url",
        "lead_doctype",
        "naming_series",

        "source",
        "status",
    )
    list_display_links = ("is_enabled",)  # keeps E124 happy
    list_editable = ("base_url", "lead_doctype", "naming_series",  "source", "status")

    def has_add_permission(self, request):
        # keep it singleton-ish
        from .models import ERPSettings
        return ERPSettings.objects.count() == 0






